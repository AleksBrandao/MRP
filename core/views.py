# core/views.py
from datetime import date, timedelta
from io import BytesIO
import csv
from decimal import Decimal
from django.db.models import Prefetch, Q, F, Value, DecimalField, FloatField, ExpressionWrapper 
from django.db.models.functions import Coalesce, Round
import logging
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from rest_framework import viewsets, filters, status
from rest_framework.views import APIView

from rest_framework.decorators import api_view
from rest_framework.response import Response

from .models import Produto, BOM, OrdemProducao, ListaTecnica, BOMSublista, BOMComponente
from .serializers import (
    ProdutoSerializer,
    BOMSerializer,
    OrdemProducaoSerializer,
    ListaTecnicaSerializer,
    BOMSublistaSerializer, 
    BOMComponenteSerializer,
)

from django.utils.functional import cached_property

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
# =========================
# ViewSets
# =========================

class ComponenteViewSet(viewsets.ModelViewSet):
    # choices agora são minúsculos; usar iexact por segurança
    queryset = Produto.objects.filter(tipo__iexact="componente")
    serializer_class = ProdutoSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["codigo", "nome", "fabricante", "codigo_fabricante"]
    ordering_fields = ["codigo", "nome", "estoque", "lead_time"]


class ListaTecnicaViewSet(viewsets.ModelViewSet):
    queryset = ListaTecnica.objects.all()
    serializer_class = ListaTecnicaSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["codigo", "nome", "observacoes"]
    ordering_fields = ["codigo", "nome", "tipo", "criado_em"]


class ProdutoViewSet(viewsets.ModelViewSet):
    queryset = Produto.objects.all()
    serializer_class = ProdutoSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["codigo", "nome", "fabricante", "codigo_fabricante"]
    ordering_fields = ["codigo", "nome", "estoque", "lead_time"]

    def get_queryset(self):
        # garante que "lista" legado não apareça mais como produto
        return Produto.objects.exclude(tipo__iexact="lista")


class BOMViewSet(viewsets.ModelViewSet):
    queryset = BOM.objects.select_related("lista_pai", "componente").all()
    serializer_class = BOMSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = [
        "lista_pai__codigo",
        "lista_pai__nome",
        "componente__codigo",
        "componente__nome",
    ]
    ordering_fields = ["lista_pai__codigo", "componente__codigo", "quantidade"]

class BOMSublistaViewSet(viewsets.ModelViewSet):
    queryset = BOMSublista.objects.select_related("lista_pai", "sublista").all()
    serializer_class = BOMSublistaSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["lista_pai__nome", "sublista__nome"]
    ordering_fields = ["lista_pai__nome", "sublista__nome"]

    def get_queryset(self):
        qs = super().get_queryset()
        lista_pai = self.request.query_params.get("lista_pai")
        if lista_pai:
            qs = qs.filter(lista_pai_id=lista_pai)
        return qs


class BOMComponenteViewSet(viewsets.ModelViewSet):
    queryset = BOMComponente.objects.select_related("lista_pai", "componente").all()
    serializer_class = BOMComponenteSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["lista_pai__nome", "componente__nome", "comentarios"]
    ordering_fields = ["lista_pai__nome", "componente__nome", "quantidade"]

    def get_queryset(self):
        qs = super().get_queryset()
        lista_pai = self.request.query_params.get("lista_pai")
        if lista_pai:
            qs = qs.filter(lista_pai_id=lista_pai)
        componente = self.request.query_params.get("componente")  # 👈 novo
        if componente:
            qs = qs.filter(componente_id=componente)
        return qs


class OrdemProducaoViewSet(viewsets.ModelViewSet):
    queryset = OrdemProducao.objects.all().order_by('-id')
    serializer_class = OrdemProducaoSerializer


# =========================
# Helpers de compatibilidade
# =========================

def _resolver_lista_da_ordem(ordem):
    """Agora a OP já referencia diretamente a Lista Técnica."""
    return ordem.lista


# =========================
# MRP (recursivo + detalhado)
# =========================

def calcular_necessidades(lista, quantidade, necessidades, nivel=0, codigo_pai=None):
    """
    Expande a BOM a partir de uma ListaTecnica (lista_pai).
    Mantém os campos esperados no frontend.
    """
    boms = BOM.objects.filter(lista_pai=lista)
    for item in boms:
        comp = item.componente
        necessidade_total = float(quantidade) * float(item.quantidade)
        estoque_atual = float(comp.estoque or 0.0)
        necessidade_liquida = max(0.0, necessidade_total - estoque_atual)

        if comp.id not in necessidades:
            necessidades[comp.id] = {
                "codigo": comp.codigo,
                "nome": _fmt_codigo_nome(comp or item.sublista),
                "necessario": necessidade_total,
                "em_estoque": estoque_atual,
                "faltando": necessidade_liquida,
                "lead_time": comp.lead_time,
                "data_compra": "",  # será preenchido depois
                "nivel": nivel,
                "codigo_pai": codigo_pai,
                "tipo": comp.tipo,  # 👈 ADICIONE ISTO
            }
        else:
            necessidades[comp.id]["necessario"] += necessidade_total
            necessidades[comp.id]["faltando"] = max(
                0.0,
                necessidades[comp.id]["necessario"] - necessidades[comp.id]["em_estoque"],
            )

        # Recursão: se você tiver sub-listas (lista dentro de lista),
        # troque 'comp' por uma ListaTecnica filha. Caso contrário, a recursão para aqui.
        # Exemplo (se usar sub-listas por parent):
        for sub_id in ListaTecnica.objects.filter(parent=lista).values_list("id", flat=True):
            sub_lista = ListaTecnica.objects.get(id=sub_id)
            calcular_necessidades(
                sub_lista,
                necessidade_total,
                necessidades,
                nivel + 1,
                codigo_pai=lista.codigo,   # 👈 importante para o front mostrar árvore
            )


def calcular_mrp_recursivo():
    necessidades = {}
    ordens = OrdemProducao.objects.all()

    for ordem in ordens:
        lista = _resolver_lista_da_ordem(ordem)
        if not lista:
            # ordem não tem lista resolvível → pula com segurança
            continue
        calcular_necessidades(lista, ordem.quantidade, necessidades, nivel=0, codigo_pai=None)

    # definir data_compra usando a menor data de entrega das OPs
    if ordens.exists():
        menor_data_entrega = min(o.data_entrega for o in ordens)
    else:
        menor_data_entrega = date.today()

    for item in necessidades.values():
        lead = int(item.get("lead_time") or 0)
        item["data_compra"] = (menor_data_entrega - timedelta(days=lead)).isoformat()

    return list(necessidades.values())


@api_view(['GET'])
def executar_mrp(request):
    necessidades = {}
    for op in OrdemProducao.objects.select_related("lista"):
        explodir_lista(op.lista, Decimal(op.quantidade), necessidades, nivel=0, codigo_pai=op.lista.codigo)
    return Response(list(necessidades.values()))


@api_view(["GET"])
def exportar_mrp_csv(request):
    resultado = {}
    for ordem in OrdemProducao.objects.select_related("lista"):
        lista = _resolver_lista_da_ordem(ordem)
        if not lista:
            continue
        adicionar_detalhes_recursivo(
            lista_id=lista.id,
            multiplicador=ordem.quantidade,
            acumulado=resultado,
            visitados=set(),
            ordem_id=ordem.id,
            lista_final_nome=lista.nome,
        )
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="resultado_mrp.csv"'
    writer = csv.writer(response)
    writer.writerow(["Código", "Nome", "Necessidade"])
    for comp_id, comp in resultado.items():
        comp_cod  = comp.get("codigo_componente") or ""
        comp_nome = comp.get("nome_componente") or ""
        comp_str  = f"{comp_cod} - {comp_nome}".strip(" -") or "—"

        writer.writerow([comp_cod, comp_nome, comp.get("necessario", 0)])
    return response



@api_view(["GET"])
def exportar_mrp_excel(request):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    print("⚙️ Iniciando exportação MRP detalhado...")
    resultado = {}
    ordens = OrdemProducao.objects.all()
    print(f"🧾 Total de ordens de produção encontradas: {ordens.count()}")

    for ordem in ordens:
        lista = _resolver_lista_da_ordem(ordem)
        print(f"🔁 Processando OP #{ordem.id} com lista {getattr(lista, 'codigo', '?')}")

        if not lista:
            print(f"⚠️ OP #{ordem.id} não possui lista associada.")
            continue

        adicionar_detalhes_recursivo(
            lista_id=lista.id,
            multiplicador=ordem.quantidade,
            acumulado=resultado,
            visitados=set(),
            ordem_id=ordem.id,
            lista_final_nome=lista.nome,
        )

    print(f"📦 Total de componentes encontrados no resultado: {len(resultado)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "MRP Detalhado"

    headers = [
        "OP","Produto Final","Qtd OP","Qtd por Unidade","Qtd Necessária",
        "Código do Componente","Nome do Componente",   # <-- novidade
        "Data Necessidade","Em Estoque","Faltando","Saldo Estoque",
    ]
    ws.append(headers)

    ids_sem_nome = [cid for cid, c in resultado.items() if not (c.get("nome_componente") or "").strip()]
    mapa_nomes = {}
    if ids_sem_nome:
        mapa_nomes = dict(
            Produto.objects.filter(id__in=ids_sem_nome).values_list("id", "nome")
        )

    for comp_id, comp in resultado.items():                 # <— precisamos do comp_id
        estoque_disponivel = comp["em_estoque"]
        for d in comp["detalhes"]:
            faltando = max(0, d["qtd_necessaria"] - estoque_disponivel)
            saldo = estoque_disponivel - d["qtd_necessaria"]
            try:
                data = (
                    OrdemProducao.objects.get(id=int(d.get("ordem_producao")))
                    .data_entrega.strftime("%d/%m/%Y")
                )
            except:
                data = "—"

            # Fallback de nome (se nome_componente vier vazio)
            comp_cod  = (comp.get("codigo_componente") or "").strip()
            comp_nome = (comp.get("nome_componente") or "").strip() or (mapa_nomes.get(comp_id) or "").strip()
            comp_str  = f"{comp_cod} - {comp_nome}".strip(" -") or "—"

            ws.append([
                d.get("ordem_producao", "—"),
                d["produto_final"] if "produto_final" in d else "—",
                d.get("qtd_produto", "—"),
                d.get("qtd_componente_por_unidade", "—"),
                d["qtd_necessaria"],
                comp_cod,                     # <- nova coluna
                comp_nome,                    # <- nova coluna
                data,
                estoque_disponivel,
                faltando,
                saldo,
            ])

            estoque_disponivel = max(0, saldo)

    # Largura automática
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="mrp_detalhado.xlsx"'
    wb.save(response)

    logger.info("✅ Arquivo Excel gerado com sucesso!")

    return response


@api_view(["GET"])
def mrp_detalhado(request):
    resultado = {}
    ordens = OrdemProducao.objects.all()

    for ordem in ordens:
        lista = _resolver_lista_da_ordem(ordem)
        if not lista:
            continue

        # Inicializa estrutura para evitar repetição do mesmo componente
        vistos = set()

        adicionar_detalhes_recursivo(
            lista_id=lista.id,
            multiplicador=ordem.quantidade,
            acumulado=resultado,
            visitados=set(),
            ordem_id=ordem.id,
            lista_final_nome=lista.nome,
        )


    # Fallback: cria detalhe genérico se houver necessidade sem detalhes
    for item in resultado.values():
        necessidade = int(item.get("necessario", 0))
        estoque = int(item.get("em_estoque", 0))
        faltando = max(0, necessidade - estoque)
        detalhes = item.get("detalhes", [])

        if faltando > 0 and not detalhes:
            lead_time = int(item.get("lead_time") or 0)
            item["detalhes"] = [{
                "tipo": "fallback",
                "descricao": "Necessidade líquida sem origem rastreável.",
                "quantidade": faltando,
                "estoque_considerado": estoque,
                "lead_time": lead_time,
                "data_sugerida": (date.today() + timedelta(days=lead_time)).isoformat(),
                "origem": None,
                "ordem_id": None,
                "lista_id": None,
            }]

    return Response(list(resultado.values()), status=status.HTTP_200_OK)


def adicionar_detalhes_recursivo(
    lista_id,
    multiplicador,
    acumulado,
    visitados,
    ordem_id,
    lista_final_nome,
    nivel=0,
    max_niveis=50,
):
    """
    Monta 'acumulado' com detalhes por componente a partir do novo modelo.
    Protege contra ciclos e limita profundidade.
    """
    if visitados is None:
        visitados = set()
    if lista_id in visitados or nivel > max_niveis:
        return
    visitados.add(lista_id)

    # componentes diretos
    for rel in BOMComponente.objects.filter(lista_pai_id=lista_id).select_related("componente"):
        comp = rel.componente
        q = Decimal(rel.quantidade or 0)
        p = Decimal(rel.ponderacao or 0)
        qpond_unidade = (q * p) / Decimal(100) if p else q
        if qpond_unidade == 0:
            continue

        qtd_total = Decimal(multiplicador) * qpond_unidade
        comp_id = comp.id

        if comp_id not in acumulado:
            acumulado[comp_id] = {
                "codigo_componente": comp.codigo,
                "id_componente": comp.id,                 # <— NOVA
                "nome_componente": comp.nome or "",       # <— AJUSTE
                "em_estoque": Decimal(comp.estoque or 0),
                "necessario": Decimal(0),
                "faltando": Decimal(0),
                "detalhes": [],
            }

        acumulado[comp_id]["necessario"] += qtd_total
        acumulado[comp_id]["faltando"] = max(
            Decimal(0),
            acumulado[comp_id]["necessario"] - acumulado[comp_id]["em_estoque"],
        )
        acumulado[comp_id]["detalhes"].append({
            "ordem_producao": ordem_id,
            "produto_final": lista_final_nome,
            "qtd_produto": multiplicador,
            "qtd_componente_por_unidade": qpond_unidade,
            "qtd_necessaria": qtd_total,
        })

    # sublistas
    for rel in BOMSublista.objects.filter(lista_pai_id=lista_id).select_related("sublista"):
        sub = rel.sublista
        if not sub:
            continue
        adicionar_detalhes_recursivo(
            lista_id=sub.id,
            multiplicador=multiplicador,
            acumulado=acumulado,
            visitados=visitados,
            ordem_id=ordem_id,
            lista_final_nome=lista_final_nome,
            nivel=nivel + 1,
            max_niveis=max_niveis,
        )

    visitados.remove(lista_id)

@api_view(["POST"])
def criar_lista_tecnica(request):
    serializer = ListaTecnicaSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def historico_produto(request, produto_id: int):
    # Histórico detalhado de um produto/componente específico
    try:
        p = Produto.objects.get(id=produto_id)
    except Produto.DoesNotExist:
        return Response({"detail": "Produto não encontrado"}, status=status.HTTP_404_NOT_FOUND)

    registros = []
    # Requer simple_history (já presente no model)
    for h in p.history.order_by("-history_date"):
        registros.append({
            "data": h.history_date.isoformat(),
            "usuario": getattr(h.history_user, "username", None),
            "acao": {"+": "Criado", "~": "Alterado", "-": "Excluído"}.get(h.history_type, h.history_type),
            "codigo": h.codigo,
            "nome": h.nome,
            "fabricante": h.fabricante,
            "codigo_fabricante": h.codigo_fabricante,
            "unidade": h.unidade,
            "estoque": float(h.estoque or 0),
            "lead_time": int(h.lead_time or 0),
            "tipo": h.tipo,
        })
    return Response(registros, status=status.HTTP_200_OK)

@api_view(["GET"])
def historico_todos_os_produtos(request):
    # Última alteração por produto (visão geral)
    out = []
    for p in Produto.objects.all().order_by("codigo"):
        h = p.history.order_by("-history_date").first()
        out.append({
            "id": p.id,
            "codigo": p.codigo,
            "nome": p.nome,
            "ultima_acao": ({"+" : "Criado","~":"Alterado","-":"Excluído"}.get(h.history_type, h.history_type) if h else None),
            "ultima_data": (h.history_date.isoformat() if h else None),
        })
    return Response(out, status=status.HTTP_200_OK)

def explodir_lista(
    lista,
    quantidade_base,
    necessidades,
    nivel=0,
    codigo_pai=None,
    visitados=None,
    max_niveis=50,
):
    """
    Agrega necessidades por componente a partir do novo modelo.
    Protege contra ciclos com 'visitados' e limita profundidade.
    """
    # guardas contra ciclo/profundidade
    if visitados is None:
        visitados = set()
    if lista.id in visitados:
        # ciclo detectado (A -> ... -> A). Ignora para não travar.
        return
    if nivel > max_niveis:
        # proteção extra
        return

    visitados.add(lista.id)
    hoje = date.today()

    # ---- 1) Componentes diretos
    comps = BOMComponente.objects.filter(lista_pai=lista).select_related("componente")
    for item in comps:
        q = Decimal(item.quantidade or 0)
        p = Decimal(item.ponderacao or 0)
        qpond_unidade = (q * p) / Decimal(100) if p else q  # se p=0, usa q cru (opcional)

        if qpond_unidade == 0:
            continue

        comp = item.componente
        comp_id = comp.id
        necessario = qpond_unidade * Decimal(quantidade_base)

        em_estoque = Decimal(comp.estoque or 0)
        faltando = max(Decimal(0), necessario - em_estoque)
        lead_time = int(getattr(comp, "lead_time", 0) or 0)
        data_compra = (hoje + timedelta(days=lead_time)).isoformat() if faltando > 0 else ""

        if comp_id not in necessidades:
            necessidades[comp_id] = {
                "id": comp_id,
                "codigo": comp.codigo,
                "nome": comp.nome,
                "necessario": Decimal(0),
                "em_estoque": em_estoque,
                "faltando": Decimal(0),
                "lead_time": lead_time,
                "data_compra": data_compra,
                "nivel": nivel,
                "codigo_pai": getattr(lista, "codigo", None),
                "tipo": getattr(comp, "tipo", ""),
            }

        necessidades[comp_id]["necessario"] += necessario
        necessidades[comp_id]["faltando"] = max(
            Decimal(0),
            necessidades[comp_id]["necessario"] - necessidades[comp_id]["em_estoque"],
        )
        if not necessidades[comp_id]["data_compra"] and necessidades[comp_id]["faltando"] > 0:
            necessidades[comp_id]["data_compra"] = data_compra

    # ---- 2) Sublistas (recursão)
    for vinc in BOMSublista.objects.filter(lista_pai=lista).select_related("sublista"):
        sub = vinc.sublista
        if not sub:
            continue
        # desce mantendo a mesma quantidade_base; passa o mesmo set (push/pop)
        explodir_lista(
            sub,
            Decimal(quantidade_base),
            necessidades,
            nivel=nivel + 1,
            codigo_pai=getattr(lista, "codigo", None),
            visitados=visitados,
            max_niveis=max_niveis,
        )

    # backtrack
    visitados.remove(lista.id)



# --- helper para montar "[CODIGO] NOME" com segurança ---
def _fmt_codigo_nome(obj):
    """
    Formata como "[CODIGO] NOME" quando houver código.
    Se não houver código, retorna apenas o nome.
    """
    if not obj:
        return ""
    codigo = (getattr(obj, "codigo", "") or "").strip()
    nome = (getattr(obj, "nome", "") or "").strip()
    return f"[{codigo}] {nome}".strip() if codigo else nome


def _cadeia_desde_raiz(no):
    """
    Retorna a cadeia de nós da RAIZ até 'no' (incluindo 'no').
    Usa o atributo 'parent' se existir em ListaTecnica. Se não existir,
    a cadeia terá apenas o próprio nó.
    """
    if not no:
        return []

    cadeia = []
    atual = no
    safety = 0
    # sobe enquanto houver parent, protegendo contra ciclos
    while atual and safety < 20:
        cadeia.append(atual)
        atual = getattr(atual, "parent", None)  # funciona mesmo se 'parent' não existir
        safety += 1

    # cadeia está do nó atual para cima; invertendo fica raiz -> ... -> nó
    return list(reversed(cadeia))

# --- helper para obter cadeia hierárquica a partir de uma lista (se existir parent) ---
def _hierarquia(lista):
    """
    Retorna até 5 níveis: [Série, Sistema, Conjunto, Subconjunto, Item]
    Se o seu modelo ListaTecnica não tiver `parent`, nada quebra: os níveis extras ficam vazios.
    """
    niveis = ["", "", "", "", ""]
    if not lista:
        return niveis

    # Caminha para cima se houver `parent`. Mantém robusto se não existir.
    cadeia = []
    atual = lista
    safety = 0
    while atual and safety < 10:
        cadeia.append(atual)
        atual = getattr(atual, "parent", None)  # se não existir parent, vira None
        safety += 1

    cadeia = list(reversed(cadeia))  # do mais alto ao mais baixo
    for i in range(min(5, len(cadeia))):
        niveis[i] = _fmt_codigo_nome(cadeia[i])

    return niveis

def _codigo_nome(obj):
    """
    Retorna (codigo, nome) de um objeto com attrs 'codigo' e 'nome'.
    Se obj for None, retorna ("", "").
    """
    if not obj:
        return "", ""
    return (getattr(obj, "codigo", "") or "").strip(), (getattr(obj, "nome", "") or "").strip()

# --- BOM (Planilha) unificado: lê BOMComponente e (opcional) BOMSublista ---

class BOMFlatView(APIView):
    """
    GET /api/bom-flat/?lista_id=...&search=...&incluir_grupos=1
    Retorna linhas em formato de planilha com níveis separados e componente (código/nome).
    Passa a consolidar:
      - BOMComponente (linhas de componente)
      - BOMSublista   (linhas de grupo, só se incluir_grupos=1)
    """

    def get(self, request, *args, **kwargs):
        from .models import BOMComponente, BOMSublista

        lista_id = request.GET.get("lista_id")
        search = (request.GET.get("search") or "").strip()
        incluir_grupos = (request.GET.get("incluir_grupos") or "").lower() in ("1", "true", "t", "yes")

        # --- Componentes (linhas "de peça") ---
        q_comp = BOMComponente.objects.select_related("lista_pai", "componente")
        if lista_id:
            q_comp = q_comp.filter(lista_pai_id=lista_id)
        if search:
            q_comp = q_comp.filter(
                Q(lista_pai__codigo__icontains=search) |
                Q(lista_pai__nome__icontains=search)   |
                Q(componente__codigo__icontains=search)|
                Q(componente__nome__icontains=search)  |
                Q(comentarios__icontains=search) |
                Q(tipo_revisao__icontains=search)
            )

        linhas: list[dict] = []

        # Monta linhas a partir do nó de referência = lista_pai
        for item in q_comp.order_by("lista_pai__codigo", "id"):
            cadeia = _cadeia_desde_raiz(item.lista_pai)  # RAIZ -> ... -> nó atual

            # extrai nomes por nível
            nomes = ["", "", "", "", ""]
            for i, nodo in enumerate(cadeia[:5]):
                _, nome = _codigo_nome(nodo)
                nomes[i] = nome
            serie_nome, sistema_nome, conjunto_nome, subconjunto_nome, item_nome = nomes

            # quantidade ponderada (seu model usa 'ponderacao')
            q = float(item.quantidade or 0)
            ponderacao = float(item.ponderacao or 0)
            quant_pond = q * (ponderacao / 100.0)

            comp_cod, comp_nom = _codigo_nome(item.componente)

            linhas.append({
                "serie_nome": serie_nome,
                "sistema_nome": sistema_nome,
                "conjunto_nome": conjunto_nome,
                "subconjunto_nome": subconjunto_nome,
                "item_nome": item_nome,  # só aparece se o nó for ITEM
                "componente_codigo": comp_cod,
                "componente_nome": comp_nom,
                "quantidade": q,
                "ponderacao": ponderacao,
                "quant_ponderada": quant_pond,
                "comentarios": item.comentarios or "",
                "tipo_revisao": item.tipo_revisao or "",      # 👈 adicionado

                    # 👇 ADIÇÕES para ação no frontend
                "linha_tipo": "componente",
                "linha_id": item.id,
                "lista_pai_id": getattr(item.lista_pai, "id", None),
            })

        # --- Grupos (sublistas) — só quando solicitado ---
        if incluir_grupos:
            q_grp = BOMSublista.objects.select_related("lista_pai", "sublista")
            if lista_id:
                q_grp = q_grp.filter(lista_pai_id=lista_id)
            if search:
                q_grp = q_grp.filter(
                    Q(lista_pai__codigo__icontains=search) |
                    Q(lista_pai__nome__icontains=search)   |
                    Q(sublista__codigo__icontains=search)  |
                    Q(sublista__nome__icontains=search)
                )

            for item in q_grp.order_by("lista_pai__codigo", "id"):
                # nó de referência (como você fazia antes): o grupo apontado
                no_ref = item.sublista or item.lista_pai
                cadeia = _cadeia_desde_raiz(no_ref)

                nomes = ["", "", "", "", ""]
                for i, nodo in enumerate(cadeia[:5]):
                    _, nome = _codigo_nome(nodo)
                    nomes[i] = nome
                serie_nome, sistema_nome, conjunto_nome, subconjunto_nome, item_nome = nomes

                linhas.append({
                    "serie_nome": serie_nome,
                    "sistema_nome": sistema_nome,
                    "conjunto_nome": conjunto_nome,
                    "subconjunto_nome": subconjunto_nome,
                    "item_nome": item_nome,
                    "componente_codigo": None,   # grupo não tem componente
                    "componente_nome": None,
                    "quantidade": None,
                    "ponderacao": None,
                    "quant_ponderada": None,
                    "comentarios": "",           # opcional
                    "tipo_revisao": "",      # 👈 adicionado

                     # 👇 ADIÇÕES para ação no frontend
                    "linha_tipo": "sublista",
                    "linha_id": item.id,  # id do BOMSublista
                    "lista_pai_id": getattr(item.lista_pai, "id", None),
                    "sublista_id": getattr(item.sublista, "id", None),
                })

        return Response(linhas, status=status.HTTP_200_OK)


class BOMFlatXLSXView(APIView):
    """
    Exporta a mesma visão (componentes + grupos quando detalhado=1) em XLSX.
    """
    def get(self, request, *args, **kwargs):
        from .models import BOMComponente, BOMSublista
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        lista_id = request.GET.get("lista_id")
        search = (request.GET.get("search") or "").strip()
        detalhado = (request.GET.get("detalhado") or "").lower() in ("1", "true", "t", "yes")

        wb = Workbook()
        ws = wb.active
        ws.title = "BOM (Planilha)"

        ws.append([
            "Série", "Sistema", "Conjunto", "Subconjunto", "Item",
            "Código do Componente", "Nome do Componente",
            "Quantidade", "Ponderação (%)", "Quant. Ponderada", "Comentários", "Tipo Revisão" # 👈 adicionado
        ])

        # Componentes
        q_comp = BOMComponente.objects.select_related("lista_pai", "componente")
        if lista_id:
            q_comp = q_comp.filter(lista_pai_id=lista_id)
        if search:
            q_comp = q_comp.filter(
                Q(lista_pai__codigo__icontains=search) |
                Q(lista_pai__nome__icontains=search)   |
                Q(componente__codigo__icontains=search)|
                Q(componente__nome__icontains=search)  |
                Q(comentarios__icontains=search) |
                Q(tipo_revisao__icontains=search)
            )

        for item in q_comp.order_by("lista_pai__codigo", "id"):
            cadeia = _cadeia_desde_raiz(item.lista_pai)
            nomes = ["", "", "", "", ""]
            for i, nodo in enumerate(cadeia[:5]):
                _, nome = _codigo_nome(nodo)
                nomes[i] = nome
            serie_nome, sistema_nome, conjunto_nome, subconjunto_nome, item_nome = nomes

            q = float(item.quantidade or 0)
            ponderacao = float(item.ponderacao or 0)
            quant_pond = q * (ponderacao / 100.0)
            comp_cod, comp_nom = _codigo_nome(item.componente)

            ws.append([
                serie_nome, sistema_nome, conjunto_nome, subconjunto_nome, item_nome,
                comp_cod, comp_nom,
                q, ponderacao, quant_pond, item.comentarios or "",
                item.tipo_revisao or "",               # 👈 novo valor
            ])

        # Grupos (quando detalhado)
        if detalhado:
            q_grp = BOMSublista.objects.select_related("lista_pai", "sublista")
            if lista_id:
                q_grp = q_grp.filter(lista_pai_id=lista_id)
            if search:
                q_grp = q_grp.filter(
                    Q(lista_pai__codigo__icontains=search) |
                    Q(lista_pai__nome__icontains=search)   |
                    Q(sublista__codigo__icontains=search)  |
                    Q(sublista__nome__icontains=search)
                )

            for item in q_grp.order_by("lista_pai__codigo", "id"):
                no_ref = item.sublista or item.lista_pai
                cadeia = _cadeia_desde_raiz(no_ref)
                nomes = ["", "", "", "", ""]
                for i, nodo in enumerate(cadeia[:5]):
                    _, nome = _codigo_nome(nodo)
                    nomes[i] = nome
                serie_nome, sistema_nome, conjunto_nome, subconjunto_nome, item_nome = nomes

                ws.append([
                    serie_nome, sistema_nome, conjunto_nome, subconjunto_nome, item_nome,
                    "", "", "", "", "", ""  # sem componente/quantidade
                ])

        # Ajuste automático de largura (limitado)
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 80)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="bom_planilha.xlsx"'
        wb.save(response)
        return response

# JSON
class BOMPlanilhaJSON(APIView):
    """
    GET /api/exports/bom/planilha.json?lista_id=...&search=...&limit=...&offset=...
    - Usa lista_pai (FK) e ponderacao (número: 10 = 10%)
    - Calcula quant_ponderada em tempo real: quantidade * (ponderacao/100) se houver ponderacao; senão, quantidade
    """
    def get(self, request):
        lista_id = request.GET.get("lista_id")
        search   = (request.GET.get("search") or "").strip()
        limit    = int(request.GET.get("limit") or 5000)
        offset   = int(request.GET.get("offset") or 0)

        qs = (BOMComponente.objects
              .select_related("lista_pai", "componente"))

        if lista_id:
            qs = qs.filter(lista_pai_id=lista_id)

        if search:
            qs = qs.filter(
                componente__nome__icontains=search
            ) | qs.filter(
                componente__codigo__icontains=search
            ) | qs.filter(
                lista_pai__nome__icontains=search
            )

        # quant_ponderada = quantidade * (Coalesce(ponderacao,0)/100) se ponderacao > 0; senão, quantidade
        # Implementação: sempre computamos quantidade * (ponderacao/100); se ponderacao nula/0, resultado é 0,
        # então somamos um termo alternativo para quando ponderacao=0: quantidade * (1 - step),
        # mas isso complica no SQL. Mais simples: use Coalesce e trate como:
        #   qp = quantidade * (Coalesce(ponderacao, 0) / 100.0)
        # e se ponderacao == 0, cai para 0. Para manter a lógica "sem ponderação => usar quantidade",
        # calculamos: qp = quantidade * (Coalesce(NULLIF(ponderacao,0), 100) / 100)
        # Como não temos NULLIF direto, faremos via Case/When.

        from django.db.models import Case, When, FloatField

        qp_expr = Case(
            When(ponderacao__gt=0,
                 then=ExpressionWrapper(
                     F("quantidade") * (Coalesce(F("ponderacao"), Value(0.0)) / Value(100.0)),
                     output_field=FloatField()
                 )),
            default=F("quantidade"),
            output_field=FloatField()
        )

        qs = qs.annotate(
            quant_pond_calc=Round(qp_expr, 2),
        )

        rows = []
        for item in qs[offset:offset+limit]:
            comp  = item.componente
            lista = item.lista_pai
            rows.append({
                "lista_id":            lista.id if lista else None,
                "lista_nome":          lista.nome if lista else "",
                "componente_id":       comp.id if comp else None,
                "componente_nome":     getattr(comp, "nome", "") or "",
                "codigo":              getattr(comp, "codigo", "") or "",
                "fabricante":          getattr(comp, "fabricante", "") or "",
                "codigo_fabricante":   getattr(comp, "codigo_fabricante", "") or "",
                "unidade":             getattr(comp, "unidade", "") or "",
                "quantidade":          float(item.quantidade or 0),
                "ponderacao":          float(getattr(item, "ponderacao", 0) or 0),  # seu campo real
                "tipo_revisao":        getattr(item, "tipo_revisao", "") or "",
                "comentarios":         getattr(item, "comentarios", "") or "",
                "quant_ponderada":     float(getattr(item, "quant_pond_calc", 0) or 0),  # 2 casas
            })

        resp = JsonResponse(rows, safe=False)
        resp["Cache-Control"] = "no-store"
        return resp


# CSV
class BOMPlanilhaCSV(APIView):
    """
    GET /api/exports/bom/planilha.csv?lista_id=...&search=...
    """
    def get(self, request):
        lista_id = request.GET.get("lista_id")
        search   = (request.GET.get("search") or "").strip()

        qs = (BOMComponente.objects
              .select_related("lista_pai", "componente"))

        if lista_id:
            qs = qs.filter(lista_pai_id=lista_id)

        if search:
            qs = qs.filter(
                componente__nome__icontains=search
            ) | qs.filter(
                componente__codigo__icontains=search
            ) | qs.filter(
                lista_pai__nome__icontains=search
            )

        def row_iter():
            header = [
                "lista_id","lista_nome","componente_id","componente_nome",
                "codigo","fabricante","codigo_fabricante","unidade",
                "quantidade","ponderacao","tipo_revisao","comentarios","quant_ponderada"
            ]
            yield ",".join(header) + "\n"

            for item in qs.iterator():
                comp  = item.componente
                lista = item.lista_pai
                quantidade = float(item.quantidade or 0)
                ponder     = float(getattr(item, "ponderacao", 0) or 0)

                # mesma regra: se ponderacao > 0, usa ponderada; senão, usa quantidade
                if ponder > 0:
                    qp = quantidade * (ponder / 100.0)
                else:
                    qp = quantidade

                row = [
                    str(lista.id if lista else ""),
                    (lista.nome if lista else "").replace(",", " "),
                    str(comp.id if comp else ""),
                    (getattr(comp, "nome", "") or "").replace(",", " "),
                    (getattr(comp, "codigo", "") or "").replace(",", " "),
                    (getattr(comp, "fabricante", "") or "").replace(",", " "),
                    (getattr(comp, "codigo_fabricante", "") or "").replace(",", " "),
                    (getattr(comp, "unidade", "") or "").replace(",", " "),
                    f"{quantidade:.2f}".replace(".", ","),   # 👈 quantidade com vírgula
                    f"{ponder:.2f}".replace(".", ","),       # 👈 ponderacao com vírgula
                    (getattr(item, "tipo_revisao", "") or "").replace(",", " "),
                    (getattr(item, "comentarios", "") or "").replace(",", " "),
                    f"{qp:.2f}".replace(".", ","),
                ]
                yield ",".join(row) + "\n"

        resp = StreamingHttpResponse(row_iter(), content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = 'inline; filename="bom_planilha.csv"'
        resp["Cache-Control"] = "no-store"
        return resp
